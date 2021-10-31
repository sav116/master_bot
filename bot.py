import io
import threading
import requests
import telebot
from telebot import types
import sqlite3
import time
import openpyxl

import config
import keyboard

from datetime import datetime

bot = telebot.TeleBot(config.token)

wb = None


# def update_google_doc(param=None):
#     while True:
#         global wb
#         response=requests.get(
#             "https://docs.google.com/spreadsheets/d/e/2PACX-1vTysMODR55FGpx2G1S6nfFxVcFQb90pysFa_LOSCjtKWeoP5lSkIo0wD4VPQ6J9FtNoX4ZOWsmQMfzV/pub?output=xlsx",
#             stream=True)
#         wb=openpyxl.load_workbook(filename=io.BytesIO(response.content), data_only=True)
#         time.sleep(300)
#
# _update_doc=threading.Thread(target=update_google_doc)
# _update_doc.start()
#
# def force_update_google_doc():
#     while True:
#         global wb
#         response=requests.get(
#             "https://docs.google.com/spreadsheets/d/e/2PACX-1vTysMODR55FGpx2G1S6nfFxVcFQb90pysFa_LOSCjtKWeoP5lSkIo0wD4VPQ6J9FtNoX4ZOWsmQMfzV/pub?output=xlsx",
#             stream=True)
#         wb=openpyxl.load_workbook(filename=io.BytesIO(response.content), data_only=True)
#         return "Google Doc обновлен!"

def update_google_doc(param=None):
    while True:
        global wb
        response = requests.get(
            "https://docs.google.com/spreadsheets/d/e/2PACX-1vRVDJabzTpcOkIVsAmmCQ5xIrGbZvun7ax0c3u8tjkcpqsCEZtmihXF_aHGGkd1U0KyJjNsO-bgY89c/pub?output=xlsx",
            stream=True)
        wb = openpyxl.load_workbook(filename=io.BytesIO(response.content), data_only=True)
        time.sleep(300)


_update_doc = threading.Thread(target=update_google_doc)
_update_doc.start()


def force_update_google_doc():
    while True:
        global wb
        response = requests.get(
            "https://docs.google.com/spreadsheets/d/e/2PACX-1vRVDJabzTpcOkIVsAmmCQ5xIrGbZvun7ax0c3u8tjkcpqsCEZtmihXF_aHGGkd1U0KyJjNsO-bgY89c/pub?output=xlsx",
            stream=True)
        wb = openpyxl.load_workbook(filename=io.BytesIO(response.content), data_only=True)
        return "Google Doc обновлен!"


@bot.message_handler(commands=['start'])
def send_welcome(message):
    userid = message.chat.id
    connect = sqlite3.connect('bot.db')
    q = connect.cursor()
    q.execute(""" CREATE TABLE IF NOT EXISTS client(
        id TEXT, surname TEXT, name1 TEXT,phone TEXT, cashback INTEGER, colvo INTEGER, adress TEXT
    ) """)
    q.execute("""CREATE TABLE IF NOT EXISTS problem(
        id TEXT, info TEXT
    )""")
    q.execute('''CREATE TABLE IF NOT EXISTS master_phone(
        id TEXT, phone TEXT
        )''')
    q.execute('''CREATE TABLE IF NOT EXISTS adm(
        id TEXT
        )''')
    connect.commit()
    res = q.execute("SELECT * FROM client where id is " + str(userid)).fetchone()
    if res is None:
        bot.send_message(message.chat.id,
                         "<b>Немного о нас:\nBlack-Izi - это Сервисный центр по ремонту телефонов, планшетов, ноутбуков.\nМы осуществляем ремонт любой сложности.\nВ сферу наших услуг в том числе входят переклейка оригинальных модульных дисплеев по заводским технологиям, BGA пайка, разблокировка iCloud, Mi, Гугл аккаунтов.\nОпыт работы наших мастеров в данной сфере более 10 лет. На все виды работ даём гарантии от 1 месяца.\nРаботать с нами вдвойне выгодно, вам также начисляются cash back баллы.</b>",
                         parse_mode='html', reply_markup=keyboard.reg)
    else:
        bot.send_message(userid, f"Пиветствую {message.from_user.username} в нашем боте", reply_markup=keyboard.profile)


@bot.message_handler(commands=['admin'])
def admin_menu(message):
    adm = []
    connect = sqlite3.connect('bot.db')
    q = connect.cursor()
    res = q.execute("SELECT * FROM adm").fetchall()
    for i in res:
        adm.append(i[0])
    if message.chat.id in config.admins or str(message.chat.id) in adm:
        bot.send_message(message.chat.id, "Админ панель", reply_markup=keyboard.admin)
    else:
        bot.send_message(message.chat.id, "Для вас эта команда недоступна")


@bot.message_handler(content_types=['text'])
def text_menu(message):
    if message.text == "Пользователи👤":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        conenct = sqlite3.connect('bot.db')
        q = conenct.cursor()
        res = q.execute("SELECT * FROM client").fetchall()
        all_users = "Все зарегистрированые пользователи\n\n"
        for i in res:
            user = f"id - {i[0]}: Фамилия - {i[1]}: Имя - {i[2]}: Телефон - {i[3]}: Кэшбек - {i[5]}: Количество заказов {i[5]}: Адрес - {i[6]}"
            all_users = all_users + str(user) + "\n\n"
        bot.send_message(message.chat.id, all_users, reply_markup=keyboard.delete)

    if message.text == "Мой профиль👤":
        try:
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            userid = str(message.chat.id)
            conenct = sqlite3.connect('bot.db')
            q = conenct.cursor()
            res = q.execute(f"SELECT * FROM client where id = {userid}").fetchone()
            name = res[2]
            phone = res[3]
            cashback = res[4]
            colvo = res[5]
            adress = res[6]
            bot.send_message(userid, f"{name} это ваш профиль👨‍💻\n\n" \
                                     f"📲Ваш номер телефона: +{phone}\n" \
                                     f"💲Кэшбек: {cashback} руб\n" \
                                     f"⚙️Количество заказов: {colvo}\n" \
                                     f"🏠Адрес вашей мастерсой: {adress}", reply_markup=keyboard.delete)
        except Exception as e:
            bot.send_message(config.coder, "❗️Произошла ошибка ПРОФИЛЬ 60❗️\n" \
                                           f"{e}")
    if message.text == "Установить номер📞":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        send = bot.send_message(message.chat.id, "Введите номер☎️")
        bot.clear_step_handler_by_chat_id(message.chat.id)
        bot.register_next_step_handler(send, setphone)

    if message.text == "Назначить админа🦸":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        send = bot.send_message(message.chat.id, "Введите id будущего админа🧞")
        bot.clear_step_handler_by_chat_id(message.chat.id)
        bot.register_next_step_handler(send, setadmin)

    if message.text == "Вернуться в меню👨‍💻":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        bot.send_message(message.chat.id, "Главное меню", reply_markup=keyboard.profile)

    if message.text == "Связаться с мастером🧑‍🔧":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        connect = sqlite3.connect('bot.db')
        q = connect.cursor()
        res = q.execute("SELECT * FROM master_phone").fetchone()
        phone = res[1]
        bot.send_message(message.chat.id, "Есть какие-то вопросы⁉️\n" \
                                          f"Обращайтесь к нашему мастеру🧑‍🔧 {phone}", reply_markup=keyboard.delete)

    if message.text == "Сделать заказ⚙️":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        bot.send_message(message.chat.id, "Выберите способ отправки устройства", reply_markup=keyboard.dostavka)

    if message.text == "Узнать цену💵":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        bot.send_message(message.chat.id, "Выберите производителя устройства📲", reply_markup=keyboard.choice_brand)
        # bot.clear_step_handler_by_chat_id(message.chat.id)
        # bot.register_next_step_handler(send, send_price)

    if message.text == "Вызвать курьера🏎":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        now = datetime.now()
        work = now.replace(hour=13, minute=0, second=0, microsecond=0)
        now_date = datetime.now()
        if work >= now_date:
            bot.send_message(message.chat.id, "Выбирите устройство", reply_markup=keyboard.devices)

            delivery = "Курьер"
        else:
            bot.send_message(message.chat.id, "Отправка курьером возможна только до 13:00")

    if message.text in config.cura:
        config.delivery[message.chat.id] = message.text

    if message.text in config.sam:
        config.delivery[message.chat.id] = message.text

    if message.text == "Списать балы🙍":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        send = bot.send_message(message.chat.id, "Введите айди пользователя которому нужно списать балы")
        bot.clear_step_handler_by_chat_id(message.chat.id)
        bot.register_next_step_handler(send, min_cashback)

    if message.text == "Приеду в мастерскую🔧":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        bot.send_message(message.chat.id, "Выбирите устройство", reply_markup=keyboard.devices)

    if message.text == "Начислить cashback💰":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        send = bot.send_message(message.chat.id, "Введите айди пользователя которому нужно начислить cashback")
        bot.clear_step_handler_by_chat_id(message.chat.id)
        bot.register_next_step_handler(send, add_cashback)
    if message.text == "Обновить Google Doc♻️":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        bot.send_message(message.chat.id, text=force_update_google_doc())
    if message.text == "Google Doc link":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        bot.send_message(message.chat.id,
                         "https://docs.google.com/spreadsheets/d/1cerKLVOwcbSgs6UQLJBUysapv_b2ACi6TlvHzxKhZJ8")


@bot.callback_query_handler(func=lambda call: True)
def answer(call):
    print(call.data)
    if call.data == "apple":
        bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)
        bot.send_message(chat_id=call.message.chat.id, text="Выберите модель", parse_mode="html",
                         reply_markup=keyboard.apple_buttons)

    elif "iPhone" in call.data or "ipad" in call.data or "watch" in call.data:
        if not call.data.startswith('send_'):
            brand = "apple"
            bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)
            bot.send_message(chat_id=call.message.chat.id, text=get_price(brand, call.data), parse_mode="html")

    # if call.data == "huawei_honor":
    #     bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)
    #     bot.send_message(chat_id=call.message.chat.id, text="Выберите модель", parse_mode="html",
    #                      reply_markup=keyboard.huawei_buttons)
    #
    # elif "Honor" in call.data or "Huawei" in call.data:
    #     if not call.data.startswith('send_'):
    #         brand = "huawei"
    #         bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)
    #         bot.send_message(chat_id=call.message.chat.id, text=get_price(brand, call.data), parse_mode="html")
    #
    # if call.data == "samsung":
    #     bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)
    #     bot.send_message(chat_id=call.message.chat.id, text="Выберите модель", parse_mode="html",
    #                      reply_markup=keyboard.samsung_buttons)
    #
    # elif "samsung" in call.data.lower():
    #     if not call.data.startswith('send_'):
    #         brand = "samsung"
    #         bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)
    #         bot.send_message(chat_id=call.message.chat.id, text=get_price(brand, call.data), parse_mode="html")
    #
    # if call.data == "xiaomi":
    #     bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)
    #     bot.send_message(chat_id=call.message.chat.id, text="Выберите модель", parse_mode="html",
    #                      reply_markup=keyboard.xiaomi_buttons)
    #
    # elif "xiaomi" in call.data.lower():
    #     if not call.data.startswith('send_'):
    #         brand = "xiaomi"
    #         bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)
    #         bot.send_message(chat_id=call.message.chat.id, text=get_price(brand, call.data), parse_mode="html")
    if call.data == "huawei_honor":
        send = bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                     text="Введите модель устройства Huawei/Honor или кодовое название", parse_mode="html")
        bot.clear_step_handler_by_chat_id(call.message.chat.id)
        bot.register_next_step_handler(send, huawei_honor_prices)

    if call.data == "samsung":
        send = bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                     text="Введите модель устройства Samsung или кодовое название", parse_mode="html")
        bot.clear_step_handler_by_chat_id(call.message.chat.id)
        bot.register_next_step_handler(send, samsung_prices)

    if call.data == "xiaomi":
        send = bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                     text="Введите модель устройства Xiaomi или кодовое название", parse_mode="html")
        bot.clear_step_handler_by_chat_id(call.message.chat.id)
        bot.register_next_step_handler(send, xiaomi_prices)

    if call.data == "oppo":
        send = bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                     text="Введите модель устройства Oppo или кодовое название", parse_mode="html")
        bot.clear_step_handler_by_chat_id(call.message.chat.id)
        bot.register_next_step_handler(send, oppo_prices)

    if call.data == "realmi":
        send = bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                     text="Введите модель устройства Realmi или кодовое название", parse_mode="html")
        bot.clear_step_handler_by_chat_id(call.message.chat.id)
        bot.register_next_step_handler(send, realmi_prices)

    if call.data == "vivo":
        send = bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                     text="Введите модель устройства Vivo или кодовое название", parse_mode="html")
        bot.clear_step_handler_by_chat_id(call.message.chat.id)
        bot.register_next_step_handler(send, vivo_prices)

    if call.data == "reg":
        send = bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                     text="Введите свое ФИО", parse_mode="html")
        bot.clear_step_handler_by_chat_id(call.message.chat.id)
        bot.register_next_step_handler(send, register)

    if call.data == "delete":
        bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)

    if call.data == "phone":
        send = bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                     text="Введите количество", parse_mode="html")
        bot.clear_step_handler_by_chat_id(call.message.chat.id)
        bot.register_next_step_handler(send, next_step)

    if call.data == "tablet":
        send = bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                     text="Введите количество", parse_mode="html")
        bot.clear_step_handler_by_chat_id(call.message.chat.id)
        bot.register_next_step_handler(send, next_step)

    if call.data == "laptop":
        send = bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                     text="Введите количество", parse_mode="html")
        bot.clear_step_handler_by_chat_id(call.message.chat.id)
        bot.register_next_step_handler(send, diagnostic1)

    if call.data == "tv":
        send = bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                     text="Введите количество", parse_mode="html")
        bot.clear_step_handler_by_chat_id(call.message.chat.id)
        bot.register_next_step_handler(send, diagnostic1)

    arr = call.data.split("_")
    if arr[0] == "send":
        try:

            connect = sqlite3.connect("bot.db")
            q = connect.cursor()
            res = q.execute(f"SELECT * FROM client where id = {call.message.chat.id}").fetchone()
            problem = q.execute(f"SELECT info FROM problem where id = {call.message.chat.id}").fetchone()[0]
            surname = res[1]
            name = res[2]
            phone = res[3]
            adress = res[6]

            bot.send_message(config.chat, "<b>❗️❗️Информация о заказе❗️❗️</b>\n\n" \
                                          f"👤ФИО: {surname} {name}\n"
                                          f"📞Номер телефона: +{phone}\n" \
                                          f"🌆Адрес мастерской: {adress}\n" \
                                          f"📲Модель: {arr[1]}\n" \
                                          f"🛒Количество: {arr[2]}\n" \
                                          f"📜Проблема: {problem}\n" \
                                          f"📦Доставка: {config.delivery[call.message.chat.id]}", parse_mode="html")
            bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)
            bot.send_message(res[0], "Ваш заказ отправлен 📤")
            q.execute(f"DELETE FROM problem where id = {call.message.chat.id}")
            connect.commit()

        except Exception as e:
            bot.send_message(config.coder, "❗️Произошла ошибка ЗАЯВКА 144❗️\n" \
                                           f"{e}")


def diagnostic1(message):
    if message.text == "Вернуться в меню👨‍💻":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        bot.send_message(message.chat.id, "Главное меню", reply_markup=keyboard.profile)
        bot.clear_step_handler_by_chat_id(message.chat.id)
    else:
        send = bot.send_message(message.chat.id, "Введите модель")
        bot.register_next_step_handler(send, diagnostic2)


def diagnostic2(message):
    if message.text == "Вернуться в меню👨‍💻":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        bot.send_message(message.chat.id, "Главное меню", reply_markup=keyboard.profile)
        bot.clear_step_handler_by_chat_id(message.chat.id)
    else:
        send = bot.send_message(message.chat.id, "Опишите вашу проблему")
        bot.register_next_step_handler(send, diagnostic3)


def diagnostic3(message):
    if message.text == "Вернуться в меню👨‍💻":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        bot.send_message(message.chat.id, "Главное меню", reply_markup=keyboard.profile)
        bot.clear_step_handler_by_chat_id(message.chat.id)

    else:
        bot.send_message(message.chat.id,
                         "<b>Для более детальной информации по вашей модели, свяжитесь с нашим мастером</b>",
                         parse_mode='html')


def setadmin(message):
    adm_id = message.text
    if message.text == "Вернуться в меню👨‍💻":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        bot.send_message(message.chat.id, "Главное меню", reply_markup=keyboard.profile)
        bot.clear_step_handler_by_chat_id(message.chat.id)

    elif adm_id.isdigit():
        try:
            connect = sqlite3.connect('bot.db')
            q = connect.cursor()
            res = q.execute(f"SELECT * FROM adm where id = {adm_id}").fetchone()
            if res is None:
                q.execute("INSERT INTO adm(id) VALUES ('%s')" % (adm_id))
                connect.commit()
                bot.send_message(message.chat.id, "Админ был успешно добавлен✅")

            bot.send_message(message.chat.id, "Админ уже добавлен😌")
        except:
            bot.send_message(message.chat.id, "Ошибка❗️")



    else:
        bot.send_message(message.chat.id, "Не корректное id❗️")

def prepare_model_string(model_string):
    model_name = ""
    code_name = ""
    code_name_start = False
    for i in model_string:
        if i == "(":
            code_name_start = True
            continue
        if i == ")":
            code_name_start = False
            continue
        if code_name_start:
            code_name += i
        if not code_name_start:
            model_name+= i
    return model_name, code_name

def code_name_is_perhaps(text):
    digit = False
    letter = False
    for i in text:
        if i.isdigit():
            digit = True
        if i.isalpha():
            letter = True
        if digit and letter:
            return True
    return False

def model_in_model_name(model, model_name):
    print(model_name)
    print(model)
    model_name_ = model_name.split()
    for i in model_name_:
        if i == model:
            return True
    return False

def vivo_prices(message):
    brand = 'vivo'
    model = message.text.lower()
    ws = wb[brand]
    result = ""
    for column in range(1, 150):
        cell_model_value = ws.cell(row=1, column=column).value
        if isinstance(cell_model_value, str):
            if model in cell_model_value.strip().lower():
                if result == "":
                    result += cell_model_value.strip() + ":\n"
                    for row in range(2, 50):
                        cell_price_value = str(ws.cell(row=row, column=column).value)
                        if cell_price_value != "None":
                            result += cell_price_value + '\n'
                else:
                    result += "\n* * * * * * * * * * * * * * * * * * * * * *\n\n"
                    result += cell_model_value.strip() + ":\n"
                    for row in range(2, 50):
                        cell_price_value = str(ws.cell(row=row, column=column).value)
                        if cell_price_value != "None":
                            result += cell_price_value + '\n'
    if result == "":
        bot.send_message(message.chat.id, "Модель не найдена!")
    else:
        bot.send_message(message.chat.id, result)

def realmi_prices(message):
    brand = 'realmi'
    model = message.text.lower()
    ws = wb[brand]
    result = ""
    for column in range(1, 150):
        cell_model_value = ws.cell(row=1, column=column).value
        if isinstance(cell_model_value, str):
            if model in cell_model_value.strip().lower():
                if result == "":
                    result += cell_model_value.strip() + ":\n"
                    for row in range(2, 50):
                        cell_price_value = str(ws.cell(row=row, column=column).value)
                        if cell_price_value != "None":
                            result += cell_price_value + '\n'
                else:
                    result += "\n* * * * * * * * * * * * * * * * * * * * * *\n\n"
                    result += cell_model_value.strip() + ":\n"
                    for row in range(2, 50):
                        cell_price_value = str(ws.cell(row=row, column=column).value)
                        if cell_price_value != "None":
                            result += cell_price_value + '\n'
    if result == "":
        bot.send_message(message.chat.id, "Модель не найдена!")
    else:
        bot.send_message(message.chat.id, result)

def oppo_prices(message):
    brand = 'oppo'
    model = message.text.lower()
    ws = wb[brand]
    result = ""
    for column in range(1, 150):
        cell_model_value = ws.cell(row=1, column=column).value
        if isinstance(cell_model_value, str):
            if model in cell_model_value.strip().lower():
                if result == "":
                    result += cell_model_value.strip() + ":\n"
                    for row in range(2, 50):
                        cell_price_value = str(ws.cell(row=row, column=column).value)
                        if cell_price_value != "None":
                            result += cell_price_value + '\n'
                else:
                    result += "\n* * * * * * * * * * * * * * * * * * * * * *\n\n"
                    result += cell_model_value.strip() + ":\n"
                    for row in range(2, 50):
                        cell_price_value = str(ws.cell(row=row, column=column).value)
                        if cell_price_value != "None":
                            result += cell_price_value + '\n'
    if result == "":
        bot.send_message(message.chat.id, "Модель не найдена!")
    else:
        bot.send_message(message.chat.id, result)

def xiaomi_prices(message):
    brand = 'xiaomi'
    model = message.text.lower()
    ws = wb[brand]
    result = ""
    for column in range(1, 150):
        cell_model_value = ws.cell(row=1, column=column).value
        if isinstance(cell_model_value, str):
            if model in cell_model_value.strip().lower():
                if result == "":
                    result += cell_model_value.strip() + ":\n"
                    for row in range(2, 50):
                        cell_price_value = str(ws.cell(row=row, column=column).value)
                        if cell_price_value != "None":
                            result += cell_price_value + '\n'
                else:
                    result += "\n* * * * * * * * * * * * * * * * * * * * * *\n\n"
                    result += cell_model_value.strip() + ":\n"
                    for row in range(2, 50):
                        cell_price_value = str(ws.cell(row=row, column=column).value)
                        if cell_price_value != "None":
                            result += cell_price_value + '\n'
    if result == "":
        bot.send_message(message.chat.id, "Модель не найдена!")
    else:
        bot.send_message(message.chat.id, result)

def samsung_prices(message):
    brand = 'samsung'
    model = message.text.lower()
    ws = wb[brand]
    result = ""
    for column in range(1, 150):
        cell_model_value = ws.cell(row=1, column=column).value
        if isinstance(cell_model_value, str):
            if model in cell_model_value.strip().lower():
                if result == "":
                    result += cell_model_value.strip() + ":\n"
                    for row in range(2, 50):
                        cell_price_value = str(ws.cell(row=row, column=column).value)
                        if cell_price_value != "None":
                            result += cell_price_value + '\n'
                else:
                    result += "\n* * * * * * * * * * * * * * * * * * * * * *\n\n"
                    result += cell_model_value.strip() + ":\n"
                    for row in range(2, 50):
                        cell_price_value = str(ws.cell(row=row, column=column).value)
                        if cell_price_value != "None":
                            result += cell_price_value + '\n'
    if result == "":
        bot.send_message(message.chat.id, "Модель не найдена!")
    else:
        bot.send_message(message.chat.id, result)

def huawei_honor_prices(message):
    brand = 'huawei'
    model = message.text.lower()
    if model == '10':
        model='honor 10 f'
    if model == '20':
        model='honor 20 '
    if model == '30':
        model='honor 30 '
    ws = wb[brand]
    code_name_is_perhaps_ = code_name_is_perhaps(model)
    result = ""
    for column in range(1, 150):
        cell_model_value = ws.cell(row=1, column=column).value
        if isinstance(cell_model_value, str):
            # print(cell_model_value.strip().lower())
            model_name, codename = prepare_model_string(cell_model_value.strip().lower())
            #if model_in_model_name(model, model_name) or (code_name_is_perhaps_ and model in codename):
            print(model_name)
            if model in model_name or (code_name_is_perhaps_ and model in codename):
                print(cell_model_value.strip().lower())
                if result == "":
                    result += cell_model_value.strip() + ":\n"
                    for row in range(2, 50):
                        cell_price_value = str(ws.cell(row=row, column=column).value)
                        if cell_price_value != "None":
                            result += cell_price_value + '\n'
                else:
                    result += "\n* * * * * * * * * * * * * * * * * * * * * *\n\n"
                    result += cell_model_value.strip() + ":\n"
                for row in range(2, 50):
                    cell_price_value = str(ws.cell(row=row, column=column).value)
                    if cell_price_value != "None":
                        result += cell_price_value + '\n'
    if result == "":
        bot.send_message(message.chat.id, "Модель не найдена!")
    else:
        bot.send_message(message.chat.id, result)


def next_step(message):
    colvo = message.text
    if message.text == "Вернуться в меню👨‍💻":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        bot.send_message(message.chat.id, "Главное меню", reply_markup=keyboard.profile)
        bot.clear_step_handler_by_chat_id(message.chat.id)

    elif colvo.isdigit():
        if int(colvo) <= 1:
            send = bot.send_message(message.chat.id, "Введите модель")
            bot.clear_step_handler_by_chat_id(message.chat.id)
            bot.register_next_step_handler(send, next_step1, colvo)
        else:
            send = bot.send_message(message.chat.id, "Введите модели ваших устройств:\n" \
                                                     "модель 1, модель 2 ...\n")
            bot.clear_step_handler_by_chat_id(message.chat.id)
            bot.register_next_step_handler(send, next_step1, colvo)

    else:
        send = bot.send_message(message.chat.id, "Введите корректное количество")
        bot.clear_step_handler_by_chat_id(message.chat.id)
        bot.register_next_step_handler(send, next_step)


def next_step1(message, colvo):
    if message.text == "Вернуться в меню👨‍💻":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        bot.send_message(message.chat.id, "Главное меню", reply_markup=keyboard.profile)
        bot.clear_step_handler_by_chat_id(message.chat.id)

    elif int(colvo) <= 1:
        model = message.text
        send = bot.send_message(message.chat.id, "Опишите вашу проблему")
        bot.clear_step_handler_by_chat_id(message.chat.id)
        bot.register_next_step_handler(send, next_step2, colvo, model)
    else:
        model = message.text
        send = bot.send_message(message.chat.id, "Опишите ваши проблемы:\n" \
                                                 "проблема 1, проблема 2 ... \n")
        bot.clear_step_handler_by_chat_id(message.chat.id)
        bot.register_next_step_handler(send, next_step2, colvo, model)


def next_step2(message, colvo, model):
    if message.text == "Вернуться в меню👨‍💻":
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        bot.send_message(message.chat.id, "Главное меню", reply_markup=keyboard.profile)
        bot.clear_step_handler_by_chat_id(message.chat.id)
    else:
        try:
            userid = message.chat.id
            conect = sqlite3.connect('bot.db')
            q = conect.cursor()
            res = q.execute(f"SELECT * FROM client where id = {userid}").fetchone()
            surname = res[1]
            name = res[2]
            phone = res[3]
            adress = res[6]
            colvo_zakazov = res[5]
            new_colvo = int(colvo_zakazov) + 1
            q.execute(f"update client set colvo = {new_colvo} where id = {userid}")
            conect.commit()
            problem = message.text
            q.execute("INSERT INTO problem(id, info) VALUES ('%s', '%s')" % (userid, problem))
            conect.commit()

            key = types.InlineKeyboardMarkup()
            send = types.InlineKeyboardButton("👍Отправить", callback_data="send_{}_{}".format(model, colvo))
            cancel = types.InlineKeyboardButton("❌Отменить", callback_data="delete")
            key.row(send, cancel)
            bot.send_message(userid, f"❗️❗️Информация о заказе❗️❗️\n\n" \
                                     f"👤ФИО: {surname} {name}\n"
                                     f"📞Номер телефона: +{phone}\n" \
                                     f"🌆Адрес мастерской: {adress}\n" \
                                     f"📲Модель: {model}\n" \
                                     f"🛒Количество: {colvo}\n" \
                                     f"📜Проблема: {problem}", parse_mode='html', reply_markup=key)



        except Exception as e:
            bot.send_message(config.coder, "❗️Произошла ошибка ЗАЯВКА 144❗️\n" \
                                           f"{e}")


def register(message):
    try:
        keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True, one_time_keyboard=True)
        button_phone = types.KeyboardButton(text="Отправить телефон📞", request_contact=True)
        keyboard.add(button_phone)
        bot.send_message(message.chat.id, 'Номер телефона', reply_markup=keyboard)
        userid = message.chat.id
        connect = sqlite3.connect('bot.db')
        q = connect.cursor()
        res = q.execute(f"SELECT * FROM client where id is " + str(userid)).fetchone()
        if res is None:
            data = message.text.split(" ")
            surname = str(data[0])
            name = str(data[1])
            cashback = 0
            colvo = 0
            q.execute("INSERT INTO client(id, surname,name1, cashback,colvo) VALUES ('%s','%s','%s','%s','%s')" % (
                userid, surname, name, cashback, colvo))
            connect.commit()
    except e:
        send = bot.send_message(message.chat.id, "Данные введены не корректно\n" \
                                                 "<b>Формат ввода: Петров Виктор Александрович</b>", parse_mode='html')
        bot.clear_step_handler_by_chat_id(call.message.chat.id)
        bot.register_next_step_handler(send, register)


@bot.message_handler(content_types=['contact'])
def contact(message):
    try:
        userid = message.chat.id
        if message.contact is not None:
            phone = message.contact.phone_number
            connect = sqlite3.connect('bot.db')
            q = connect.cursor()
            res = q.execute(f"SELECT * FROM client where id = '{userid}'").fetchone()
            if res[3] is None:
                q.execute(f"update client set phone = {phone} where id = {userid}")
                connect.commit()
            send = bot.send_message(message.chat.id, "Введите адрес своей мастерской")
            bot.clear_step_handler_by_chat_id(message.chat.id)
            bot.register_next_step_handler(send, register1)
    except Exception as e:
        bot.send_message(config.coder, "❗️Произошла ошибка КОНТАКТЫ❗️\n" \
                                       f"{e}")


def register1(message):
    try:
        userid = message.chat.id
        adress = str(message.text)
        connect = sqlite3.connect('bot.db')
        q = connect.cursor()
        res = q.execute(f"SELECT * FROM client where id = '{userid}'").fetchone()
        if res[6] is None:
            q.execute(f"update client set adress = '{adress}' where id = {userid}")
            connect.commit()
            bot.send_message(userid, "<b>Регистрация прошла успешно</b>", parse_mode='html',
                             reply_markup=keyboard.profile)
    except Exception as e:
        bot.send_message(config.coder, "❗️Произошла ошибка РЕГИСТРАЦИЯ❗️\n" \
                                       f"{e}")


def add_cashback(message):
    chat_id = message.text
    send = bot.send_message(message.chat.id, "Введите сумму кэшека для пользователя")
    bot.clear_reply_handlers_by_message_id(message.chat.id)
    bot.register_next_step_handler(send, add_money, chat_id)


def add_money(message, chat_id):
    summ = message.text
    if summ.isdigit():
        try:
            summ = int(summ)
            connect = sqlite3.connect('bot.db')
            q = connect.cursor()
            balik = q.execute(f"SELECT cashback FROM client where id = {chat_id}").fetchone()[0]
            balik = int(balik + summ)
            q.execute(f"update client set cashback = {balik} where id = {chat_id}")
            connect.commit()
            bot.send_message(message.chat.id, "Начисления прошло успешно✅")
            bot.send_message(chat_id, f"Вам начислен кэшбек в размере {summ} рублей")
        except Exception as e:
            bot.send_message(message.chat.id, f"Произошла ошибка")

            bot.send_message(config.coder, "❗️Произошла ошибка НАЧИСЛЕНИЯ❗️\n" \
                                           f"{e}")
    else:
        bot.send_message(message.chat.id, "Введите корректные данные")


def min_cashback(message):
    chat_id = message.text
    connect = sqlite3.connect('bot.db')
    q = connect.cursor()
    bal = q.execute(f"SELECT cashback FROM client where id = {chat_id}").fetchone()[0]
    send = bot.send_message(message.chat.id, f"Введите сумму списания.\nДоступно {bal}руб")
    bot.clear_reply_handlers_by_message_id(message.chat.id)
    bot.register_next_step_handler(send, min_money, chat_id, bal)


def min_money(message, chat_id, bal):
    summ = message.text
    if summ.isdigit() and int(summ) <= int(bal):
        new_summ = int(bal) - int(summ)
        connect = sqlite3.connect('bot.db')
        q = connect.cursor()
        q.execute(f"update client set cashback = {new_summ} where id = {chat_id}")
        connect.commit()
        bot.send_message(message.chat.id, "Балы успешно списаны✅")
        bot.send_message(chat_id, f"У вас были списанны балы на сумму {summ}")
    else:
        bot.send_message(message.chat.id, "Не корректная сумма списания👿")


def setphone(message):
    try:
        new_phone = message.text
        user_id = message.chat.id
        connect = sqlite3.connect('bot.db')
        q = connect.cursor()
        res = q.execute("SELECT * FROM master_phone").fetchone()
        if res is None:
            q.execute("INSERT INTO master_phone (id, phone) VALUES ('%s', '%s')" % (user_id, new_phone))
            connect.commit()
        else:
            q.execute(f"update master_phone set phone = {new_phone} where id = {user_id}")
            connect.commit()

        bot.send_message(user_id, "Номер телефона установлен успешно")
    except:
        bot.send_message(user_id, "Произошла ошибка")


def get_price(brand, model):
    ws = wb[brand]
    result = ""
    for column in range(1, 30):
        cell_model_value = ws.cell(row=1, column=column).value
        if isinstance(cell_model_value, str):
            if model.lower() in cell_model_value.strip().lower():
                result += cell_model_value.strip() + ":\n\n"
                for row in range(2, 50):
                    cell_price_value = str(ws.cell(row=row, column=column).value)
                    if cell_price_value != "None":
                        result += cell_price_value + '\n'
                return result
    return "Не могу найти цены на устройство"


while True:
    try:
        bot.polling(none_stop=True, interval=0)
    except Exception as e:
        print(e)

        time.sleep(15)
